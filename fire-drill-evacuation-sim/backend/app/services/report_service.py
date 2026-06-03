import io
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from sqlalchemy.orm import Session

from app.models.models import DrillReport, Building, PeopleGroup, RescueVehicle, EventLog, FireIncident


class ReportService:

    def generate_report(self, db: Session, drill_id: int) -> Optional[DrillReport]:
        report = db.query(DrillReport).filter(DrillReport.id == drill_id).first()
        if not report:
            return None

        total_buildings = db.query(Building).count()
        total_people = sum(pg.count for pg in db.query(PeopleGroup).all())
        evacuated_people = sum(pg.count for pg in db.query(PeopleGroup).filter(PeopleGroup.status == "evacuated").all())
        total_vehicles = db.query(RescueVehicle).count()

        events = db.query(EventLog).order_by(EventLog.timestamp).all()
        fire_events = [e for e in events if e.event_type in ("fire_detected", "fire_contained", "fire_extinguished")]

        fire_containment_time = None
        fire_detected = None
        fire_contained = None
        for e in fire_events:
            if e.event_type == "fire_detected" and not fire_detected:
                fire_detected = e.timestamp
            if e.event_type == "fire_contained" and not fire_contained:
                fire_contained = e.timestamp
        if fire_detected and fire_contained:
            fire_containment_time = (fire_contained - fire_detected).total_seconds() / 60.0

        evacuation_events = [e for e in events if e.event_type in ("evacuation_started", "drill_end")]
        avg_evacuation_time = None
        if len(evacuation_events) >= 2:
            start = evacuation_events[0].timestamp
            end = evacuation_events[-1].timestamp
            avg_evacuation_time = (end - start).total_seconds() / 60.0

        active_fires = db.query(FireIncident).filter(FireIncident.status == "active").count()
        contained_fires = db.query(FireIncident).filter(FireIncident.status == "contained").count()
        extinguished_fires = db.query(FireIncident).filter(FireIncident.status == "extinguished").count()

        building_statuses = {}
        for b in db.query(Building).all():
            building_statuses[b.status] = building_statuses.get(b.status, 0) + 1

        vehicle_statuses = {}
        for v in db.query(RescueVehicle).all():
            vehicle_statuses[v.status] = vehicle_statuses.get(v.status, 0) + 1

        statistics = {
            "fires": {"active": active_fires, "contained": contained_fires, "extinguished": extinguished_fires},
            "building_statuses": building_statuses,
            "vehicle_statuses": vehicle_statuses,
            "total_events": len(events),
            "evacuation_rate": round(evacuated_people / total_people * 100, 1) if total_people > 0 else 0,
        }

        report.total_buildings = total_buildings
        report.total_people = total_people
        report.evacuated_people = evacuated_people
        report.total_vehicles = total_vehicles
        report.average_evacuation_time = avg_evacuation_time
        report.fire_containment_time = fire_containment_time
        report.statistics = statistics
        report.summary = (
            f"演练共涉及{total_buildings}栋建筑、{total_people}人，"
            f"成功疏散{evacuated_people}人，"
            f"疏散率{statistics['evacuation_rate']}%。"
            f"共出动{total_vehicles}辆救援车辆。"
        )

        db.commit()
        db.refresh(report)
        return report

    def export_pdf(self, db: Session, report_id: int) -> bytes:
        report = db.query(DrillReport).filter(DrillReport.id == report_id).first()
        if not report:
            return b""

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)

        elements = []

        elements.append(Spacer(1, 10 * mm))

        title_data = [[f"Fire Drill Report - {report.drill_name}"]]
        title_table = Table(title_data, colWidths=[180 * mm])
        title_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTSIZE', (0, 0), (-1, -1), 16),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#1a5276')),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ]))
        elements.append(title_table)
        elements.append(Spacer(1, 8 * mm))

        info_data = [
            ["Drill Name", report.drill_name],
            ["Start Time", str(report.start_time)],
            ["End Time", str(report.end_time) if report.end_time else "N/A"],
            ["Total Buildings", str(report.total_buildings)],
            ["Total People", str(report.total_people)],
            ["Evacuated People", str(report.evacuated_people)],
            ["Total Vehicles", str(report.total_vehicles)],
            ["Avg Evacuation Time (min)", str(round(report.average_evacuation_time, 2)) if report.average_evacuation_time else "N/A"],
            ["Fire Containment Time (min)", str(round(report.fire_containment_time, 2)) if report.fire_containment_time else "N/A"],
        ]

        info_table = Table(info_data, colWidths=[70 * mm, 110 * mm])
        info_table.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#eaf2f8')),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 6 * mm))

        if report.summary:
            summary_data = [["Summary"], [report.summary]]
            summary_table = Table(summary_data, colWidths=[180 * mm])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a5276')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            elements.append(summary_table)

        doc.build(elements)
        buffer.seek(0)
        return buffer.read()

    def export_excel(self, db: Session, report_id: int) -> bytes:
        report = db.query(DrillReport).filter(DrillReport.id == report_id).first()
        if not report:
            return b""

        wb = Workbook()
        ws = wb.active
        ws.title = "Drill Report"

        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="1A5276", end_color="1A5276", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        ws.merge_cells('A1:D1')
        ws['A1'] = f"Fire Drill Report - {report.drill_name}"
        ws['A1'].font = Font(bold=True, size=16, color="1A5276")
        ws['A1'].alignment = Alignment(horizontal="center")

        headers = ["Item", "Value"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        data_rows = [
            ("Drill Name", report.drill_name),
            ("Start Time", str(report.start_time)),
            ("End Time", str(report.end_time) if report.end_time else "N/A"),
            ("Total Buildings", report.total_buildings),
            ("Total People", report.total_people),
            ("Evacuated People", report.evacuated_people),
            ("Total Vehicles", report.total_vehicles),
            ("Avg Evacuation Time (min)", round(report.average_evacuation_time, 2) if report.average_evacuation_time else "N/A"),
            ("Fire Containment Time (min)", round(report.fire_containment_time, 2) if report.fire_containment_time else "N/A"),
            ("Summary", report.summary or "N/A"),
        ]

        for row_idx, (item, value) in enumerate(data_rows, 4):
            ws.cell(row=row_idx, column=1, value=item).border = thin_border
            ws.cell(row=row_idx, column=2, value=value).border = thin_border

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 50

        if report.statistics:
            ws2 = wb.create_sheet("Statistics")
            stat_headers = ["Category", "Key", "Value"]
            for col, header in enumerate(stat_headers, 1):
                cell = ws2.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            row = 2
            for category, values in report.statistics.items():
                if isinstance(values, dict):
                    for key, val in values.items():
                        ws2.cell(row=row, column=1, value=category).border = thin_border
                        ws2.cell(row=row, column=2, value=str(key)).border = thin_border
                        ws2.cell(row=row, column=3, value=str(val)).border = thin_border
                        row += 1
                else:
                    ws2.cell(row=row, column=1, value=category).border = thin_border
                    ws2.cell(row=row, column=2, value="-").border = thin_border
                    ws2.cell(row=row, column=3, value=str(values)).border = thin_border
                    row += 1

            ws2.column_dimensions['A'].width = 20
            ws2.column_dimensions['B'].width = 20
            ws2.column_dimensions['C'].width = 20

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.read()
